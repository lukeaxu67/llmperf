from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from pydantic import BaseModel, Field
import pandas as pd
import openpyxl  # noqa: F401
from openpyxl.utils import get_column_letter

from ..base_analysis import BaseAnalysis
from ..record_query import RecordQuery
from ..analysis_registry import register_analysis
from ..utils import _params_key, _safe_filename, _group_key, _percentile

from llmperf.records.model import RunRecord, field_aliases


def _summary_for_records(items: List[RunRecord]) -> Dict[str, Any]:
    first_resp = [float(r.first_resp_time) for r in items if r.first_resp_time]
    char_per_sec = [float(r.char_per_second) for r in items if r.char_per_second]
    token_throughput = [float(r.token_throughput) for r in items if r.token_throughput]
    total_cost = sum(float(r.total_cost) for r in items)
    count = float(len(items))
    success_count = float(len([r for r in items if int(r.status) == 200]))
    error_rate = 0.0 if count <= 0 else float((count - success_count) / count)
    currency = items[0].currency if items else ""
    return {
        "count": count,
        "success_count": success_count,
        "error_rate": error_rate,
        "avg_first_resp_time": (sum(first_resp) / len(first_resp)) if first_resp else 0.0,
        "p95_first_resp_time": _percentile(first_resp, 0.95) if first_resp else 0.0,
        "avg_char_per_second": (sum(char_per_sec) / len(char_per_sec)) if char_per_sec else 0.0,
        "avg_token_throughput": (sum(token_throughput) / len(token_throughput)) if token_throughput else 0.0,
        "total_cost": total_cost,
        "currency": currency,
    }


def _detail_dict(record: RunRecord) -> Dict[str, Any]:
    return {
        "run_id": record.run_id,
        "executor_id": record.executor_id,
        "dataset_row_id": record.dataset_row_id,
        "provider": record.provider,
        "model": record.model,
        "status": int(record.status),
        "qtokens": int(record.qtokens),
        "atokens": int(record.atokens),
        "ctokens": int(record.ctokens),
        "first_resp_time": int(record.first_resp_time),
        "last_resp_time": int(record.last_resp_time),
        "char_per_second": float(record.char_per_second),
        "token_throughput": float(record.token_throughput),
        "prompt_cost": float(record.prompt_cost),
        "completion_cost": float(record.completion_cost),
        "cache_cost": float(record.cache_cost),
        "total_cost": float(record.total_cost),
        "currency": record.currency,
        "request_params": _params_key(record.request_params),
    }


_AGGREGATE_ALIASES: Dict[str, str] = {
    "count": "请求数",
    "success_count": "成功数",
    "error_rate": "错误率",
    "avg_first_resp_time": "平均首响(ms)",
    "p95_first_resp_time": "P95首响(ms)",
    "avg_char_per_second": "平均字符速度(char/s)",
    "avg_token_throughput": "平均吞吐(tok/s)",
}


@register_analysis("summary")
class SummaryExportAnalysis(BaseAnalysis["SummaryExportAnalysis.Config"]):

    config: Config

    class Config(BaseModel):
        run_id: str = Field(min_length=1)
        task_name: str = "run"
        output_dir: str = "."
        query: RecordQuery = Field(default_factory=RecordQuery)

    def run(self) -> Dict[str, Any]:
        storage = self.config.query.storage()
        records = list(storage.fetch_run_records(self.config.run_id))

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"{_safe_filename(self.config.task_name)}-{timestamp}.xlsx"
        out_path = Path(self.config.output_dir).expanduser().resolve() / filename
        out_path.parent.mkdir(parents=True, exist_ok=True)

        grouped: Dict[Tuple[str, str, str], List[RunRecord]] = defaultdict(list)
        for rec in records:
            grouped[_group_key(rec)].append(rec)

        summary_rows: List[Dict[str, Any]] = []
        for (provider, model, params_json), items in grouped.items():
            row = {
                "provider": provider,
                "model": model,
                "request_params": params_json,
                **_summary_for_records(items),
            }
            summary_rows.append(row)

        summary_df = pd.DataFrame(summary_rows)

        detail_dfs: List[Tuple[str, "pd.DataFrame"]] = []
        for idx, ((provider, model, params_json), items) in enumerate(grouped.items(), start=1):
            sheet_name = f"{provider}.{model}.{idx}"
            detail_df = pd.DataFrame([_detail_dict(r) for r in items])
            detail_dfs.append((sheet_name, detail_df))

        aliases = dict(field_aliases(RunRecord))
        aliases.update(_AGGREGATE_ALIASES)

        def _rename_columns(df: "pd.DataFrame") -> "pd.DataFrame":
            return df.rename(columns={k: aliases.get(k, k) for k in df.columns})

        def _round_df(df: "pd.DataFrame") -> "pd.DataFrame":
            numeric_cols = df.select_dtypes(include=["float", "int"]).columns
            out = df.copy()
            for col in numeric_cols:
                out[col] = out[col].astype(float).round(2)
            return out

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            summary_out = _rename_columns(_round_df(summary_df))
            summary_out.to_excel(writer, sheet_name="汇总", index=False)

            for sheet_name, df in detail_dfs:
                df_out = _rename_columns(_round_df(df))
                safe_sheet = sheet_name[:31]
                df_out.to_excel(writer, sheet_name=safe_sheet, index=False)

            wb = writer.book
            for ws in wb.worksheets:
                max_width = 60
                for col_idx, cell in enumerate(ws[1], start=1):
                    col_letter = get_column_letter(col_idx)
                    header = str(cell.value or "")
                    width = max(10, min(max_width, len(header) * 2))
                    ws.column_dimensions[col_letter].width = width
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, float):
                            cell.number_format = "0.00"

        return {"output_path": str(out_path)}
